import sqlite3
import csv
import openpyxl
import os


def chonkyboi():
    inpsql3 = sqlite3.connect('projects.db')
    sql3_cursor = inpsql3.cursor()
    sql3_cursor.execute('SELECT * FROM project_information')
    with open('project_information.csv', 'w') as out_csv_file:
        csv_out = csv.writer(out_csv_file)
        # write header
        csv_out.writerow([d[0] for d in sql3_cursor.description])
        # write data
        for result in sql3_cursor:
            csv_out.writerow(result)

    sql3_cursor.execute('SELECT * FROM collars')
    with open('collars.csv', 'w') as out_csv_file:
        csv_out = csv.writer(out_csv_file)
        # write header
        csv_out.writerow([d[0] for d in sql3_cursor.description])
        # write data
        for result in sql3_cursor:
            csv_out.writerow(result)
    inpsql3.close()

    # merge together

    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]  # Remove the default 'Sheet1'

    for filename in ['collars.csv', 'project_information.csv']:
        with open(filename, newline='') as f_input:
            root, ext = os.path.splitext(filename)
            ws = wb.create_sheet(title=root)

            for row in csv.reader(f_input, delimiter=','):
                ws.append(row)

    wb.save('logging_data.xlsx')

    # delete csv files
    files = ['collars.csv', 'project_information.csv']
    for filename in files:
        os.remove(filename)


if __name__ == "__main__":
    chonkyboi()
